import base64
from io import BytesIO

from odoo import _, fields, models
from odoo.exceptions import UserError

UNITS_ALIASES = {
    'cbm': 'cbm', 'rt': 'rt', 'cwt': 'cwt', 'shmt': 'shmt',
    'b/l': 'bl', 'bl': 'bl', '20 dc': '20dc', '20dc': '20dc',
    '40 hc': '40hc', '40hc': '40hc',
}

LCL_COLUMNS = {'pol': 'pol', 'pod': 'pod', 'rates name': 'rates_name', 'rates': 'rates', 'units': 'units', 'section': 'section'}
FCL_COLUMNS = {'name': 'name', 'rates': 'rates', 'units': 'units', 'section': 'section'}

FCL_SECTION_ALIASES = {
    'ocean rate': 'ocean_rate', 'ocean': 'ocean_rate',
    'local charge': 'local_charge', 'local': 'local_charge',
    'exw fee': 'exw_fee', 'exw': 'exw_fee',
}
LCL_SECTION_ALIASES = {
    'lcl rate': 'lcl_rate', 'rate': 'lcl_rate',
    'lcl exw fee': 'lcl_exw_fee', 'exw fee': 'lcl_exw_fee', 'exw': 'lcl_exw_fee',
    'lcl local charge': 'lcl_local_charge', 'local charge': 'lcl_local_charge',
    'fcl local charge': 'fcl_local_charge',
}
AIR_CHARGE_SECTION_ALIASES = {
    'air local charge': 'air_local_charge', 'local charge': 'air_local_charge',
    'air exw fee': 'air_exw_fee', 'exw fee': 'air_exw_fee', 'exw': 'air_exw_fee',
}
TRUCKING_CHARGE_SECTION_ALIASES = {
    'trucking surcharge': 'trucking_surcharge', 'surcharge': 'trucking_surcharge',
    'customs': 'trucking_surcharge', 'customs fee': 'trucking_surcharge',
}
CHARGE_COLUMNS = {'name': 'name', 'rates': 'rates', 'units': 'units', 'section': 'section'}
AIR_COLUMNS = {
    'aol': 'aol', 'aod': 'aod',
    '- 45 kgs': 'kg_minus_45', '-45 kgs': 'kg_minus_45', '-45': 'kg_minus_45',
    '+ 45 kgs': 'kg_plus_45', '+45 kgs': 'kg_plus_45', '+45': 'kg_plus_45',
    '+ 100 kgs': 'kg_plus_100', '+100 kgs': 'kg_plus_100', '+100': 'kg_plus_100',
    '+ 300 kgs': 'kg_plus_300', '+300 kgs': 'kg_plus_300', '+300': 'kg_plus_300',
    '+ 500 kgs': 'kg_plus_500', '+500 kgs': 'kg_plus_500', '+500': 'kg_plus_500',
    '+ 1000 kgs': 'kg_plus_1000', '+1000 kgs': 'kg_plus_1000', '+1000': 'kg_plus_1000',
}
TRUCKING_COLUMNS = {
    'truck 1.5 tons': 'truck_1_5t', 'truck 1.5t': 'truck_1_5t',
    'truck 3.5 tons': 'truck_3_5t', 'truck 3.5t': 'truck_3_5t',
    'truck 5 tons': 'truck_5t', 'truck 5t': 'truck_5t',
    '1-3 cbms': 'cbm_1_3', '1-3 cbm': 'cbm_1_3', '1- 3 cbms': 'cbm_1_3',
    '3-5 cbms': 'cbm_3_5', '3-5 cbm': 'cbm_3_5',
    '5-10 cbms': 'cbm_5_10', '5-10 cbm': 'cbm_5_10', '5- 10 cbms': 'cbm_5_10',
    '20 dc': 'dc_20', '20dc': 'dc_20',
    '40 dc': 'dc_40', '40dc': 'dc_40',
}


class PricelistRatesImportWizard(models.TransientModel):
    _name = 'freight.pricelist.rates.import.wizard'
    _description = 'Import Rates from File into Pricelist'

    pricelist_id = fields.Many2one('master.freight.pricelist', required=True)
    service_type_code = fields.Char(related='pricelist_id.service_type_code')
    import_file = fields.Binary(string='File', required=True)
    file_name = fields.Char(string='File Name')

    def action_import_rates(self):
        self.ensure_one()
        if not self.file_name:
            raise UserError(_('Please upload a file.'))

        ext = self.file_name.rsplit('.', 1)[-1].lower() if '.' in self.file_name else ''

        if ext in ('pdf', 'png', 'jpg', 'jpeg'):
            raise UserError(_('PDF and image import via AI OCR is coming soon. Please upload an Excel file (.xlsx) for now.'))

        if ext not in ('xlsx', 'xls'):
            raise UserError(_('Unsupported file format. Please upload an Excel file (.xlsx).'))

        data = base64.b64decode(self.import_file)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
        except Exception as e:
            raise UserError(_('Could not read Excel file: %s') % str(e))

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise UserError(_('The uploaded file contains no data rows.'))

        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        data_rows = rows[1:]
        # Skip fully empty rows
        data_rows = [r for r in data_rows if any(c is not None for c in r)]

        if not data_rows:
            raise UserError(_('The uploaded file contains no data rows.'))

        service_code = self.pricelist_id.service_type_code
        if service_code == 'lcl':
            count = self._import_lcl(headers, data_rows)
        elif service_code == 'fcl':
            count = self._import_fcl(headers, data_rows)
        elif service_code == 'air':
            count = self._import_air(headers, data_rows)
            count += self._import_air_charges(headers, data_rows)
        elif service_code == 'trucking':
            count = self._import_trucking(headers, data_rows)
            count += self._import_trucking_charges(headers, data_rows)
        else:
            raise UserError(_('Please set a Service Type on the pricelist before importing.'))

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Successful'),
                'message': _('%d rate line(s) imported.') % count,
                'type': 'success',
                'next': {'type': 'ir.actions.act_window_close'},
            },
        }

    def _map_columns(self, headers, column_map):
        mapping = {}
        for idx, header in enumerate(headers):
            if header in column_map:
                mapping[column_map[header]] = idx
        return mapping

    def _resolve_port(self, value, row_num):
        if not value:
            return False
        val = str(value).strip()
        Port = self.env['master.port.airport']
        port = Port.search([('code', '=ilike', val)], limit=1)
        if not port:
            port = Port.search([('name', '=ilike', val)], limit=1)
        if not port:
            raise UserError(_('Row %d: could not find port/airport "%s".') % (row_num, val))
        return port.id

    def _normalize_units(self, value, row_num):
        if not value:
            return False
        key = UNITS_ALIASES.get(str(value).strip().lower())
        if key is None:
            raise UserError(_('Row %d: unrecognized units value "%s".') % (row_num, value))
        return key

    def _get_float(self, value):
        try:
            return float(value) if value else 0.0
        except (ValueError, TypeError):
            return 0.0

    def _resolve_section(self, value, aliases, default):
        if not value:
            return default
        key = str(value).strip().lower()
        if key in aliases:
            return aliases[key]
        # Try exact match against alias values
        if key in aliases.values():
            return key
        return default

    def _import_lcl(self, headers, data_rows):
        col = self._map_columns(headers, LCL_COLUMNS)
        self.pricelist_id.lcl_line_ids.unlink()
        # Default section based on partner type: carrier → lcl_rate, agent → lcl_exw_fee
        default_section = 'lcl_rate' if self.pricelist_id.partner_type == 'carrier' else 'lcl_exw_fee'
        vals_list = []
        for i, row in enumerate(data_rows, start=2):
            vals = {'pricelist_id': self.pricelist_id.id}
            if 'section' in col:
                vals['section'] = self._resolve_section(
                    row[col['section']], LCL_SECTION_ALIASES, default_section)
            else:
                vals['section'] = default_section
            if 'pol' in col:
                vals['pol_id'] = self._resolve_port(row[col['pol']], i)
            if 'pod' in col:
                vals['pod_id'] = self._resolve_port(row[col['pod']], i)
            if 'rates_name' in col:
                vals['rates_name'] = str(row[col['rates_name']] or '')
            if 'rates' in col:
                vals['rates'] = self._get_float(row[col['rates']])
            if 'units' in col:
                vals['units'] = self._normalize_units(row[col['units']], i)
            vals_list.append(vals)
        self.env['master.freight.pricelist.lcl'].create(vals_list)
        return len(vals_list)

    def _import_fcl(self, headers, data_rows):
        col = self._map_columns(headers, FCL_COLUMNS)
        self.pricelist_id.fcl_line_ids.unlink()
        # Default section based on partner type: carrier → ocean_rate, agent → local_charge
        default_section = 'ocean_rate' if self.pricelist_id.partner_type == 'carrier' else 'local_charge'
        vals_list = []
        for i, row in enumerate(data_rows, start=2):
            vals = {'pricelist_id': self.pricelist_id.id}
            if 'section' in col:
                vals['section'] = self._resolve_section(
                    row[col['section']], FCL_SECTION_ALIASES, default_section)
            else:
                vals['section'] = default_section
            if 'name' in col:
                vals['name'] = str(row[col['name']] or '')
            if 'rates' in col:
                vals['rates'] = self._get_float(row[col['rates']])
            if 'units' in col:
                vals['units'] = self._normalize_units(row[col['units']], i)
            vals_list.append(vals)
        self.env['master.freight.pricelist.fcl'].create(vals_list)
        return len(vals_list)

    def _import_air(self, headers, data_rows):
        col = self._map_columns(headers, AIR_COLUMNS)
        self.pricelist_id.air_line_ids.unlink()
        vals_list = []
        for i, row in enumerate(data_rows, start=2):
            vals = {'pricelist_id': self.pricelist_id.id}
            if 'aol' in col:
                vals['aol_id'] = self._resolve_port(row[col['aol']], i)
            if 'aod' in col:
                vals['aod_id'] = self._resolve_port(row[col['aod']], i)
            for field in ('kg_minus_45', 'kg_plus_45', 'kg_plus_100',
                          'kg_plus_300', 'kg_plus_500', 'kg_plus_1000'):
                if field in col:
                    vals[field] = self._get_float(row[col[field]])
            vals_list.append(vals)
        self.env['master.freight.pricelist.air'].create(vals_list)
        return len(vals_list)

    def _import_trucking(self, headers, data_rows):
        col = self._map_columns(headers, TRUCKING_COLUMNS)
        self.pricelist_id.trucking_line_ids.unlink()
        vals_list = []
        for i, row in enumerate(data_rows, start=2):
            vals = {'pricelist_id': self.pricelist_id.id}
            for field in ('truck_1_5t', 'truck_3_5t', 'truck_5t',
                          'cbm_1_3', 'cbm_3_5', 'cbm_5_10', 'dc_20', 'dc_40'):
                if field in col:
                    vals[field] = self._get_float(row[col[field]])
            vals_list.append(vals)
        self.env['master.freight.pricelist.trucking'].create(vals_list)
        return len(vals_list)

    def _import_air_charges(self, headers, data_rows):
        """Import charge lines (local charge / exw fee) for Air from a 'charges' sheet."""
        wb_data = base64.b64decode(self.import_file)
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(wb_data), read_only=True, data_only=True)
        if 'charges' not in wb.sheetnames:
            return 0
        ws = wb['charges']
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return 0
        ch_headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        ch_data = [r for r in rows[1:] if any(c is not None for c in r)]
        if not ch_data:
            return 0
        col = self._map_columns(ch_headers, CHARGE_COLUMNS)
        # Clear existing air charge lines
        self.pricelist_id.air_local_charge_ids.unlink()
        self.pricelist_id.air_exw_fee_ids.unlink()
        vals_list = []
        for i, row in enumerate(ch_data, start=2):
            section = 'air_local_charge'
            if 'section' in col:
                section = self._resolve_section(
                    row[col['section']], AIR_CHARGE_SECTION_ALIASES, 'air_local_charge')
            vals = {
                'pricelist_id': self.pricelist_id.id,
                'section': section,
            }
            if 'name' in col:
                vals['name'] = str(row[col['name']] or '')
            if 'rates' in col:
                vals['rates'] = self._get_float(row[col['rates']])
            if 'units' in col:
                vals['units'] = self._normalize_units(row[col['units']], i)
            vals_list.append(vals)
        self.env['master.freight.pricelist.charge'].create(vals_list)
        return len(vals_list)

    def _import_trucking_charges(self, headers, data_rows):
        """Import surcharge lines for Trucking from a 'charges' sheet."""
        wb_data = base64.b64decode(self.import_file)
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(wb_data), read_only=True, data_only=True)
        if 'charges' not in wb.sheetnames:
            return 0
        ws = wb['charges']
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return 0
        ch_headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        ch_data = [r for r in rows[1:] if any(c is not None for c in r)]
        if not ch_data:
            return 0
        col = self._map_columns(ch_headers, CHARGE_COLUMNS)
        # Clear existing trucking surcharge lines
        self.pricelist_id.trucking_surcharge_ids.unlink()
        vals_list = []
        for i, row in enumerate(ch_data, start=2):
            vals = {
                'pricelist_id': self.pricelist_id.id,
                'section': 'trucking_surcharge',
            }
            if 'name' in col:
                vals['name'] = str(row[col['name']] or '')
            if 'rates' in col:
                vals['rates'] = self._get_float(row[col['rates']])
            if 'units' in col:
                vals['units'] = self._normalize_units(row[col['units']], i)
            vals_list.append(vals)
        self.env['master.freight.pricelist.charge'].create(vals_list)
        return len(vals_list)
